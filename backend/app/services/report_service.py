# S.A.M.I. - Servicio de Reportes
import asyncio
import logging
import os
import json
from datetime import datetime, timedelta, date
from typing import Dict, List, Optional, Any
import pandas as pd
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

from ..core.config import settings
from ..core.database import get_db
from ..models.report import Report, ReportTemplate, ReportData, ReportSchedule

logger = logging.getLogger(__name__)

class ReportService:
    """Servicio para generación y distribución de reportes"""
    
    def __init__(self):
        self.templates = {}
        self.schedules = {}
        self.running = False
        
    async def initialize(self):
        """Inicializar el servicio de reportes"""
        try:
            # Cargar plantillas de reportes
            await self.load_report_templates()
            
            # Cargar programaciones de reportes
            await self.load_report_schedules()
            
            # Iniciar scheduler de reportes
            await self.start_report_scheduler()
            
            logger.info("Servicio de reportes inicializado correctamente")
            
        except Exception as e:
            logger.error(f"Error inicializando servicio de reportes: {e}")
            raise
    
    async def load_report_templates(self):
        """Cargar plantillas de reportes"""
        # Plantillas por defecto
        default_templates = [
            {
                "name": "Reporte Diario Operativo",
                "report_type": "daily",
                "category": "operational",
                "format": "pdf",
                "is_auto_generated": True,
                "generation_schedule": "0 6 * * *",  # 6:00 AM diario
                "template_content": self._get_daily_report_template(),
                "data_sources": ["employees", "assets", "fuel", "events"],
                "default_recipients": ["admin@empresa.com", "supervisor@empresa.com"]
            },
            {
                "name": "Reporte Semanal de Proyectos",
                "report_type": "weekly",
                "category": "projects",
                "format": "excel",
                "is_auto_generated": True,
                "generation_schedule": "0 8 * * 1",  # Lunes 8:00 AM
                "template_content": self._get_weekly_projects_template(),
                "data_sources": ["projects", "employees", "assets", "fuel"],
                "default_recipients": ["manager@empresa.com"]
            },
            {
                "name": "Reporte Mensual Financiero",
                "report_type": "monthly",
                "category": "financial",
                "format": "pdf",
                "is_auto_generated": True,
                "generation_schedule": "0 9 1 * *",  # Día 1 de cada mes 9:00 AM
                "template_content": self._get_monthly_financial_template(),
                "data_sources": ["fuel", "projects", "assets", "employees"],
                "default_recipients": ["admin@empresa.com", "accounting@empresa.com"]
            }
        ]
        
        for template in default_templates:
            self.templates[template["name"]] = template
    
    async def load_report_schedules(self):
        """Cargar programaciones de reportes"""
        # Programaciones por defecto
        default_schedules = [
            {
                "name": "Reporte Diario 6AM",
                "template_name": "Reporte Diario Operativo",
                "cron_expression": "0 6 * * *",
                "is_active": True
            },
            {
                "name": "Reporte Semanal Lunes",
                "template_name": "Reporte Semanal de Proyectos", 
                "cron_expression": "0 8 * * 1",
                "is_active": True
            },
            {
                "name": "Reporte Mensual",
                "template_name": "Reporte Mensual Financiero",
                "cron_expression": "0 9 1 * *",
                "is_active": True
            }
        ]
        
        for schedule in default_schedules:
            self.schedules[schedule["name"]] = schedule
    
    async def start_report_scheduler(self):
        """Iniciar scheduler de reportes automáticos"""
        try:
            self.running = True
            
            # Crear task para scheduler
            asyncio.create_task(self._report_scheduler_worker())
            
            logger.info("Scheduler de reportes iniciado")
            
        except Exception as e:
            logger.error(f"Error iniciando scheduler de reportes: {e}")
            raise
    
    async def _report_scheduler_worker(self):
        """Worker para scheduler de reportes"""
        while self.running:
            try:
                current_time = datetime.utcnow()
                
                # Verificar cada programación
                for schedule_name, schedule in self.schedules.items():
                    if schedule["is_active"]:
                        # Verificar si es hora de generar el reporte
                        if await self._should_generate_report(schedule, current_time):
                            await self._generate_scheduled_report(schedule)
                
                # Esperar 1 minuto antes de verificar nuevamente
                await asyncio.sleep(60)
                
            except Exception as e:
                logger.error(f"Error en scheduler de reportes: {e}")
                await asyncio.sleep(60)
    
    async def _should_generate_report(self, schedule: Dict, current_time: datetime) -> bool:
        """Verificar si es hora de generar un reporte"""
        # Implementación básica de verificación de cron
        # En producción, usar una librería como croniter
        try:
            cron_parts = schedule["cron_expression"].split()
            if len(cron_parts) != 5:
                return False
            
            minute, hour, day, month, weekday = cron_parts
            
            # Verificar minuto y hora
            if minute != "*" and int(minute) != current_time.minute:
                return False
            if hour != "*" and int(hour) != current_time.hour:
                return False
            
            # Verificar día del mes
            if day != "*" and int(day) != current_time.day:
                return False
            
            # Verificar mes
            if month != "*" and int(month) != current_time.month:
                return False
            
            # Verificar día de la semana
            if weekday != "*" and int(weekday) != current_time.weekday():
                return False
            
            return True
            
        except Exception as e:
            logger.error(f"Error verificando programación: {e}")
            return False
    
    async def _generate_scheduled_report(self, schedule: Dict):
        """Generar reporte programado"""
        try:
            template_name = schedule["template_name"]
            if template_name not in self.templates:
                logger.error(f"Plantilla no encontrada: {template_name}")
                return
            
            template = self.templates[template_name]
            
            # Generar reporte
            report_data = await self.generate_report(
                template_name=template_name,
                report_date=datetime.utcnow().date(),
                recipients=template.get("default_recipients", [])
            )
            
            logger.info(f"Reporte programado generado: {template_name}")
            
        except Exception as e:
            logger.error(f"Error generando reporte programado: {e}")
    
    async def generate_report(self, template_name: str, report_date: date, 
                            recipients: List[str] = None) -> Dict:
        """Generar reporte específico"""
        try:
            if template_name not in self.templates:
                raise ValueError(f"Plantilla no encontrada: {template_name}")
            
            template = self.templates[template_name]
            
            # Recopilar datos
            data = await self._collect_report_data(template["data_sources"], report_date)
            
            # Generar archivo
            file_path = await self._generate_report_file(template, data, report_date)
            
            # Enviar reporte
            if recipients:
                await self._send_report(file_path, recipients, template_name)
            
            return {
                "success": True,
                "template_name": template_name,
                "report_date": report_date,
                "file_path": file_path,
                "recipients": recipients or [],
                "generated_at": datetime.utcnow()
            }
            
        except Exception as e:
            logger.error(f"Error generando reporte: {e}")
            return {
                "success": False,
                "error": str(e),
                "template_name": template_name,
                "report_date": report_date
            }
    
    async def _collect_report_data(self, data_sources: List[str], report_date: date) -> Dict:
        """Recopilar datos para el reporte"""
        data = {}
        
        try:
            # Simular recopilación de datos
            # En implementación real, consultar la base de datos
            
            if "employees" in data_sources:
                data["employees"] = {
                    "total": 25,
                    "present": 20,
                    "absent": 5,
                    "overtime": 3
                }
            
            if "assets" in data_sources:
                data["assets"] = {
                    "total": 50,
                    "in_use": 35,
                    "available": 10,
                    "maintenance": 5
                }
            
            if "fuel" in data_sources:
                data["fuel"] = {
                    "total_consumed": 1250.5,
                    "total_cost": 125000.0,
                    "average_price": 100.0,
                    "tanks": [
                        {"name": "Tanque Principal", "level": 75.5, "capacity": 1000},
                        {"name": "Tanque Auxiliar", "level": 45.2, "capacity": 500}
                    ]
                }
            
            if "projects" in data_sources:
                data["projects"] = {
                    "active": 3,
                    "completed": 1,
                    "on_hold": 0,
                    "total_budget": 500000.0,
                    "actual_cost": 350000.0
                }
            
            if "events" in data_sources:
                data["events"] = {
                    "total": 15,
                    "resolved": 12,
                    "pending": 3,
                    "critical": 1
                }
            
            return data
            
        except Exception as e:
            logger.error(f"Error recopilando datos: {e}")
            return {}
    
    async def _generate_report_file(self, template: Dict, data: Dict, report_date: date) -> str:
        """Generar archivo de reporte"""
        try:
            # Crear directorio de reportes
            reports_dir = os.path.join(settings.reports_dir, str(report_date.year), str(report_date.month))
            os.makedirs(reports_dir, exist_ok=True)
            
            # Generar nombre de archivo
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"{template['name'].replace(' ', '_')}_{timestamp}.{template['format']}"
            file_path = os.path.join(reports_dir, filename)
            
            # Generar archivo según formato
            if template["format"] == "pdf":
                await self._generate_pdf_report(template, data, file_path, report_date)
            elif template["format"] == "excel":
                await self._generate_excel_report(template, data, file_path, report_date)
            else:
                raise ValueError(f"Formato no soportado: {template['format']}")
            
            return file_path
            
        except Exception as e:
            logger.error(f"Error generando archivo de reporte: {e}")
            raise
    
    async def _generate_pdf_report(self, template: Dict, data: Dict, file_path: str, report_date: date):
        """Generar reporte en formato PDF"""
        try:
            doc = SimpleDocTemplate(file_path, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=TA_CENTER
            )
            story.append(Paragraph(template["name"], title_style))
            story.append(Spacer(1, 12))
            
            # Fecha
            date_style = ParagraphStyle(
                'DateStyle',
                parent=styles['Normal'],
                fontSize=12,
                alignment=TA_RIGHT
            )
            story.append(Paragraph(f"Fecha: {report_date.strftime('%d/%m/%Y')}", date_style))
            story.append(Spacer(1, 20))
            
            # Contenido del reporte
            if template["name"] == "Reporte Diario Operativo":
                await self._add_daily_report_content(story, data, styles)
            elif template["name"] == "Reporte Semanal de Proyectos":
                await self._add_weekly_projects_content(story, data, styles)
            elif template["name"] == "Reporte Mensual Financiero":
                await self._add_monthly_financial_content(story, data, styles)
            
            # Construir PDF
            doc.build(story)
            
        except Exception as e:
            logger.error(f"Error generando PDF: {e}")
            raise
    
    async def _generate_excel_report(self, template: Dict, data: Dict, file_path: str, report_date: date):
        """Generar reporte en formato Excel"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte"
            
            # Título
            ws['A1'] = template["name"]
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Fecha
            ws['A2'] = f"Fecha: {report_date.strftime('%d/%m/%Y')}"
            ws['A2'].font = Font(size=12)
            
            # Contenido
            row = 4
            for section, section_data in data.items():
                ws[f'A{row}'] = section.upper()
                ws[f'A{row}'].font = Font(size=14, bold=True)
                row += 1
                
                if isinstance(section_data, dict):
                    for key, value in section_data.items():
                        ws[f'A{row}'] = key
                        ws[f'B{row}'] = value
                        row += 1
                
                row += 1
            
            # Ajustar columnas
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 20
            
            wb.save(file_path)
            
        except Exception as e:
            logger.error(f"Error generando Excel: {e}")
            raise
    
    async def _add_daily_report_content(self, story, data, styles):
        """Agregar contenido del reporte diario"""
        # Empleados
        if "employees" in data:
            story.append(Paragraph("RESUMEN DE PERSONAL", styles['Heading2']))
            emp_data = data["employees"]
            table_data = [
                ["Métrica", "Valor"],
                ["Total Empleados", str(emp_data["total"])],
                ["Presentes", str(emp_data["present"])],
                ["Ausentes", str(emp_data["absent"])],
                ["Horas Extra", str(emp_data["overtime"])]
            ]
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 20))
        
        # Combustible
        if "fuel" in data:
            story.append(Paragraph("RESUMEN DE COMBUSTIBLE", styles['Heading2']))
            fuel_data = data["fuel"]
            story.append(Paragraph(f"Consumo Total: {fuel_data['total_consumed']} litros", styles['Normal']))
            story.append(Paragraph(f"Costo Total: ${fuel_data['total_cost']:,.2f}", styles['Normal']))
            story.append(Paragraph(f"Precio Promedio: ${fuel_data['average_price']:,.2f}/litro", styles['Normal']))
            story.append(Spacer(1, 20))
    
    async def _add_weekly_projects_content(self, story, data, styles):
        """Agregar contenido del reporte semanal de proyectos"""
        if "projects" in data:
            story.append(Paragraph("ESTADO DE PROYECTOS", styles['Heading2']))
            proj_data = data["projects"]
            story.append(Paragraph(f"Proyectos Activos: {proj_data['active']}", styles['Normal']))
            story.append(Paragraph(f"Proyectos Completados: {proj_data['completed']}", styles['Normal']))
            story.append(Paragraph(f"Presupuesto Total: ${proj_data['total_budget']:,.2f}", styles['Normal']))
            story.append(Paragraph(f"Costo Actual: ${proj_data['actual_cost']:,.2f}", styles['Normal']))
            story.append(Spacer(1, 20))
    
    async def _add_monthly_financial_content(self, story, data, styles):
        """Agregar contenido del reporte mensual financiero"""
        if "fuel" in data:
            story.append(Paragraph("ANÁLISIS FINANCIERO MENSUAL", styles['Heading2']))
            fuel_data = data["fuel"]
            story.append(Paragraph(f"Gasto en Combustible: ${fuel_data['total_cost']:,.2f}", styles['Normal']))
            story.append(Paragraph(f"Consumo Total: {fuel_data['total_consumed']} litros", styles['Normal']))
            story.append(Spacer(1, 20))
    
    async def _send_report(self, file_path: str, recipients: List[str], subject: str):
        """Enviar reporte por email"""
        try:
            if not settings.smtp_server:
                logger.warning("Servidor SMTP no configurado, no se enviará el email")
                return
            
            # Crear mensaje
            msg = MIMEMultipart()
            msg['From'] = settings.smtp_username
            msg['To'] = ", ".join(recipients)
            msg['Subject'] = f"SAMI - {subject}"
            
            # Cuerpo del mensaje
            body = f"""
            Estimado/a,
            
            Se adjunta el reporte: {subject}
            
            Este es un reporte automático generado por el Sistema SAMI.
            
            Saludos,
            Sistema SAMI
            """
            msg.attach(MIMEText(body, 'plain'))
            
            # Adjuntar archivo
            with open(file_path, "rb") as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename= {os.path.basename(file_path)}'
                )
                msg.attach(part)
            
            # Enviar email
            server = smtplib.SMTP(settings.smtp_server, settings.smtp_port)
            server.starttls()
            server.login(settings.smtp_username, settings.smtp_password)
            text = msg.as_string()
            server.sendmail(settings.smtp_username, recipients, text)
            server.quit()
            
            logger.info(f"Reporte enviado a {len(recipients)} destinatarios")
            
        except Exception as e:
            logger.error(f"Error enviando reporte: {e}")
    
    def _get_daily_report_template(self) -> str:
        """Obtener plantilla de reporte diario"""
        return """
        <h1>Reporte Diario Operativo</h1>
        <p>Fecha: {report_date}</p>
        
        <h2>Resumen de Personal</h2>
        <p>Total Empleados: {employees_total}</p>
        <p>Presentes: {employees_present}</p>
        <p>Ausentes: {employees_absent}</p>
        
        <h2>Resumen de Combustible</h2>
        <p>Consumo Total: {fuel_consumed} litros</p>
        <p>Costo Total: ${fuel_cost}</p>
        """
    
    def _get_weekly_projects_template(self) -> str:
        """Obtener plantilla de reporte semanal de proyectos"""
        return """
        <h1>Reporte Semanal de Proyectos</h1>
        <p>Semana: {week_start} - {week_end}</p>
        
        <h2>Estado de Proyectos</h2>
        <p>Proyectos Activos: {projects_active}</p>
        <p>Proyectos Completados: {projects_completed}</p>
        <p>Presupuesto Total: ${projects_budget}</p>
        """
    
    def _get_monthly_financial_template(self) -> str:
        """Obtener plantilla de reporte mensual financiero"""
        return """
        <h1>Reporte Mensual Financiero</h1>
        <p>Mes: {month_year}</p>
        
        <h2>Análisis de Costos</h2>
        <p>Gasto en Combustible: ${fuel_cost}</p>
        <p>Gasto en Mantenimiento: ${maintenance_cost}</p>
        <p>Gasto Total: ${total_cost}</p>
        """
    
    async def get_available_templates(self) -> List[Dict]:
        """Obtener plantillas disponibles"""
        return list(self.templates.values())
    
    async def get_report_history(self, limit: int = 50) -> List[Dict]:
        """Obtener historial de reportes generados"""
        # Esta función debería consultar la base de datos
        return []
    
    async def get_system_status(self) -> Dict:
        """Obtener estado del servicio de reportes"""
        return {
            "running": self.running,
            "templates_count": len(self.templates),
            "schedules_count": len(self.schedules),
            "active_schedules": len([s for s in self.schedules.values() if s["is_active"]]),
            "last_updated": datetime.utcnow()
        }
